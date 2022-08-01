# -*- coding: utf-8 -*-
import dataiku
import pandas as pd, numpy as np
from dataiku import pandasutils as pdu

from dataiku import SQLExecutor2

import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import io 
import re

from tempfile import NamedTemporaryFile


# Write Excel to managed folder
def write_wb_to_managed_folder(wb,output_folder, file_name):
    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        output = tmp.read()
        with output_folder.get_writer(file_name) as w:
            w.write(output)

# Read Input Template as Workbook
def read_wb_from_managed_folder(input_folder):
    template_file_name = input_folder.list_paths_in_partition()[0]
    with input_folder.get_download_stream(template_file_name) as f:
        bytes_in = io.BytesIO(f.read())
        wb = openpyxl.load_workbook(bytes_in)
    return wb

# Write dataset to sheet at an Row/ Column position
def populate_table_in_ws(df, ws, start_row, start_col):
    df_np = df.to_numpy()
    for row_num in range(df.shape[0]):
        for col_num in range (df.shape[1]):
            ws.cell(row = (row_num + start_row), column = (col_num +start_col)).value =  df_np[row_num][col_num]
    return ws


# Find a SQL tag in worksheets and return coordinates
def find_tags_in_ws(ws,query_tag,row_max, col_max):
    tags = []
    for row_idx in range(1,row_max):
        for col_idx in range (1,col_max):
            value = ws.cell(row = row_idx,column =col_idx).value
            if value != None:
                if (str(value).startswith(query_tag)):
                    tags.append([value,row_idx, col_idx])
    return tags


def parse_query_from_raw_query(raw_query):
    result = re.search('<(.*)>', str(raw_query))
    if result != None:
        query = result.group(1)
    else:
        raise Exception("No query found in the tagged cell of value : {}".format(raw_query))
    return query


def get_df_from_query(cnx_name,query):
    executor = SQLExecutor2(connection=cnx_name)
    df = executor.query_to_df(query)
    return df
